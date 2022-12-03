# -*- coding: utf-8 -*-
"""

Description: check risk control items required in the quotation process to ensure compliance.

Note: Sensitive information has been obfuscated.

"""

''' quotation check'''
import os
import pandas as pd

path = "xxxxx"
excels = [
    pd.read_excel(path + fname)
    for fname in os.listdir(path)
    if ".xlsx" in fname and "keywordA" in fname
]


name = [
    fname[:-5]  
    for fname in os.listdir(path)
    if ".xlsx" in fname and "keywordA" in fname
]


value = [
    pd.read_excel(path + fname)
    for fname in os.listdir(path)
    if ".xls" in fname and "keywordB" in fname
]

transfer=pd.read_excel(path + 'transfer.xlsx')

import numpy as np
from WindPy import *
w.start()

df2 = value[0]

a_lst_name = []
for k in range(0,len(excels)):
    df1 = excels[k]
    count=df1[(df1.iloc[:,22]==1)].shape[0]+6
    #get count of rows
    
    #get data from Wind
    wind = w.wss(df1.iloc[2,0], "ipo_op_uplimit,ipo_op_downlimit,ipo_op_minofchg,ipo_inq_startdate,ipo_inq_enddate,ipo_newshares,ipo_op_startdate").Data
    wind_data = [int(wind[0][0]),int(wind[1][0]),int(wind[2][0]),str(wind[3][0])[0:-9],str(wind[4][0])[0:-9],wind[5][0]/10000,str(wind[6][0])[0:-9]]
    human_data = [df1.iloc[3,4],df1.iloc[3,2],df1.iloc[3,5],df1.iloc[3,8],df1.iloc[3,11],df1.iloc[3,13]]
    
    if(wind_data[0] != human_data[0]):
        print('Upper Limit Error')      #up limit
    if(wind_data[1] != human_data[1]):
        print('Lower Limit Error')      #low limit
    if(wind_data[2] != human_data[2]): 
        print('Step Size Error')        #step
    if(wind_data[3] != human_data[3]):
        print('Inq Start Error')        #inq start date
    if(wind_data[4] != human_data[4]):
        print('Inq End Error')          #inq end date
    if(wind_data[5] != human_data[5]):
        print('New Shares Error')       #new issue shares
    for i in range(6,count):
        if(df1.iloc[i,13] != wind_data[6]):
            print('Op Start Error in line {}'.format(i+2))         # operation date
    
    # check date
    today = df2.iloc[4:,0]    
    cnt = len(set(today))    
    if(cnt != 1):
        print('Date Error')      #date error
    time = str(df1.iloc[4,15])   #date standardized
    for i in range(4,df2.shape[0]):
        if (df2.iloc[i,0] not in time):
            print('Time Error in line {}'.format(i+2))  #
    
    # check current account
    for i in range(6,count):
        current = df1.iloc[i,18]        #current account
        payment = df1.iloc[i,21]        #payable
        if(current < 100):
            print('Current Below One Milion in line {}'.format(i+2))    #current below 1 million
        if(current < payment):
            print('Money Is Not Enough in line {}'.format(i+2))         #current account enough to pay
    
    # check amount
    price = df1.iloc[3,1]               #price
    for i in range(6,count):
        title = df1.iloc[i,1]
        amount = df1.iloc[i,3]          #amount
        if(amount < human_data[1]):
            print('Amount lower than inf in product{}'.format(title))
        if(amount > human_data[0]):
            print('Amount higher than sup in product{}'.format(title))
        pre_total = round(df1.iloc[i,5],0)       
        total = round(price * amount,0)     
        
        if(total != pre_total):         
            print('Amount Error in product{}'.format(title))

        try:
            new_title=transfer[transfer.iloc[:,0]==title].iloc[0,1]
            try:
                available=df2[(df2.iloc[:,2]==new_title)].iloc[0,7]
                if total*10000 > available:    
                    print('Out Of Range in product {}'.format(title))
            except:
                print('Net value of {} not found'.format(title))
        except:
            print('product {} not found in transfer'.format(title))                       
                         
    print("{} check finished.".format(name[k]))
    a_lst_name.append(name[k])
    
    string=string.replace(","," ")
    print(string)

#%% 

'''excel process'''

from openpyxl import *

filename = [
    fname 
    for fname in os.listdir(path)
    if ".xlsx" in fname and "keywordC" in fname and "keywordD" not in fname
]

excels = [
    pd.read_excel(path+fname)
    for fname in os.listdir(path)
    if ".xlsx" in fname and "keywordC" in fname and "keywordD" not in fname
]

def delete_rows_private(name):
    wb = load_workbook(path + name)
    ws = wb.active
    ws.delete_rows(count+3,100)
    i = 1
    max_row = count+2
    while i>0:
        i = 0
        for row in ws.iter_rows():
            if row[1].value == "FundNameA" or row[1].value == "FundNameB":
                row_number = row[1].row
                ws.delete_rows(row_number)
                i = i + 1
                max_row -= 1
                break
            else:
                continue
    global count_public        
    count_public = max_row
    wb.save(path+name)
    

def delete_rows_public(name):
    wb = load_workbook(path + name)
    ws = wb.active
    ws.delete_rows(count+3,100)
    i = 1
    max_row = count+2
    while i>0:
        i = 0
        for j in range(8,max_row):
            if ws.cell(j,2).value != "FundNameA" and ws.cell(j,2).value != "FundNameB":
                ws.delete_rows(j)
                i = i + 1
                max_row -= 1
                break
            else:
                continue
    global count_private        
    count_private = max_row
    wb.save(path+name)
