# -*- coding: utf-8 -*-
"""

Description: process raw excel files and divide each into two with specific requirements.

Note: Sensitive information has been obfuscated.

"""

import os
import pandas as pd
from openpyxl import *


filename = [
    fname 
    for fname in os.listdir(path)
    if ".xlsx" in fname and "keywordC" in fname and "keywordD" not in fname
]

excels = [
    pd.read_excel(path+fname)
    for fname in os.listdir(path)
    if ".xlsx" in fname and "keywordC" in fname and "keywordD" not in fname
]

def delete_rows_private(name):
    wb = load_workbook(path + name)
    ws = wb.active
    ws.delete_rows(count+3,100)
    i = 1
    max_row = count+2
    while i>0:
        i = 0
        for row in ws.iter_rows():
            if row[1].value == "FundNameA" or row[1].value == "FundNameB":
                row_number = row[1].row
                ws.delete_rows(row_number)
                i = i + 1
                max_row -= 1
                break
            else:
                continue
    global count_public        
    count_public = max_row
    wb.save(path+name)
    

def delete_rows_public(name):
    wb = load_workbook(path + name)
    ws = wb.active
    ws.delete_rows(count+3,100)
    i = 1
    max_row = count+2
    while i>0:
        i = 0
        for j in range(8,max_row):
            if ws.cell(j,2).value != "FundNameA" and ws.cell(j,2).value != "FundNameB":
                ws.delete_rows(j)
                i = i + 1
                max_row -= 1
                break
            else:
                continue
    global count_private        
    count_private = max_row
    wb.save(path+name)